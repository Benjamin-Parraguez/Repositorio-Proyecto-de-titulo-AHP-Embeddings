import numpy as np
import pandas as pd
from scipy.linalg import eig
from fractions import Fraction
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import itertools
import matplotlib.pyplot as plt

# === Paso 1: Leer matrices desde archivo Excel ===
archivo_excel = r"C:\Users\rikig\OneDrive\Escritorio\Proyecto de Título\Mesa 5\Metodos Multicriterios\AHP\Python\AHP_CasoNoticias.xlsx"  # <- Cambia por tu nombre de archivo
xls = pd.ExcelFile(archivo_excel)

# Función para convertir fracción en string a float
def fraccion_a_float(x):
    try:
        return float(Fraction(str(x)))
    except:
        return np.nan

# -----------------------------
# LEER PESOS DIRECTOS DE CRITERIOS
# -----------------------------
# Espera hoja "criterios" con dos columnas: criterio | peso
df_pesos = pd.read_excel(xls, sheet_name="criterios", header=0)

# Normalizar nombres de columnas
cols_lower = [str(c).strip().lower() for c in df_pesos.columns]
col_crit = df_pesos.columns[cols_lower.index("criterio")]
col_peso = df_pesos.columns[cols_lower.index("peso")]

# Lista de criterios (se usará para leer hojas de alternativas)
criterios = df_pesos[col_crit].astype(str).tolist()

# Parseo robusto de pesos con coma decimal o '%'
def _parse_num(x):
    if isinstance(x, str):
        x = x.replace('%', '').strip().replace(',', '.')
    return float(x)

pesos_criterios = df_pesos[col_peso].map(_parse_num).to_numpy(dtype=float)

# Si los pesos vienen en %, convierto; luego normalizo por seguridad
if 99.0 <= pesos_criterios.sum() <= 101.0:
    pesos_criterios = pesos_criterios / 100.0
pesos_criterios = pesos_criterios / pesos_criterios.sum()

# Guardamos esta tabla para exportarla luego (en vez de matriz de criterios)
matriz_criterios_df = df_pesos[[col_crit, col_peso]].copy()

# -----------------------------
# LEER MATRICES DE ALTERNATIVAS
# -----------------------------
matrices_alternativas = {}
for criterio in criterios:
    # OJO: el nombre de la hoja debe ser el nombre del criterio en minúsculas.
    # Ej: "tecnico", "economico", "social", "ambiental", "comercial"
    df_alt = pd.read_excel(xls, sheet_name=criterio.lower(), index_col=0)
    matrices_alternativas[criterio] = df_alt.applymap(fraccion_a_float).to_numpy()
alternativas = list(df_alt.columns)

# === Paso 2: Funciones ===
def calcular_pesos(matriz):
    valores, vectores = eig(matriz)
    max_index = np.argmax(valores.real)
    vector_principal = vectores[:, max_index].real
    pesos = vector_principal / sum(vector_principal)
    return pesos

def consistencia_ahp(matriz):
    valores, _ = eig(matriz)
    lambda_max = max(valores.real)
    n = matriz.shape[0]
    CI = (lambda_max - n) / (n - 1)
    RI_dict = {1: 0.0, 2: 0.0, 3: 0.58, 4: 0.90, 5: 1.12,
               6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
    RI = RI_dict.get(n, 1.49)
    CR = CI / RI if RI else 0
    return lambda_max, CI, CR

# === Paso 3: Cálculos de pesos ===
# Ya tenemos pesos_criterios desde Excel (no se calcula con eigenvector)
pesos_locales = [calcular_pesos(matrices_alternativas[c]) for c in criterios]

pesos_globales = sum(pesos_locales[i] * pesos_criterios[i] for i in range(len(criterios)))

# No aplica consistencia de criterios (no hay matriz)
lambda_max, CI, CR = (np.nan, np.nan, np.nan)

# === Paso 4: Crear tablas para exportar ===
df_resultados = pd.DataFrame({
    'Alternativa': alternativas,
    'Peso Global': pesos_globales
})

df_criterios = pd.DataFrame({
    'Criterio': criterios,
    'Peso': pesos_criterios
})

df_consistencia = pd.DataFrame({
    'Lambda máx': [lambda_max],
    'CI': [CI],
    'CR': [CR],
    'Consistencia aceptable': ['N/A (pesos directos)']
})

# === Paso 5: Exportar a Excel con formato ===
archivo_salida = r"C:\Users\rikig\OneDrive\Escritorio\Proyecto de Título\Mesa 5\Metodos Multicriterios\AHP\Python\R_AHP_CasoNoticias.xlsx"
with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
    df_criterios.to_excel(writer, sheet_name='Criterios', index=False)
    df_resultados.to_excel(writer, sheet_name='Resultados', index=False)
    df_consistencia.to_excel(writer, sheet_name='Consistencia', index=False)

# === Paso 6: Exportar información adicional al Excel ===
with pd.ExcelWriter(archivo_salida, engine='openpyxl', mode='a') as writer:
    # 1) Exportar "Matriz_Criterios" -> ahora es la tabla de pesos directos
    matriz_criterios_df.to_excel(writer, sheet_name='Matriz_Criterios', index=False)

    # 2) Exportar matrices de alternativas por criterio
    for criterio in criterios:
        df_alt_original = pd.read_excel(xls, sheet_name=criterio.lower(), index_col=0)
        df_alt_original.to_excel(writer, sheet_name=f'Matriz_{criterio}')

    # 3) Exportar pesos locales por criterio
    for i, criterio in enumerate(criterios):
        df_local = pd.DataFrame({
            'Alternativa': alternativas,
            f'Peso Local ({criterio})': pesos_locales[i]
        })
        df_local.to_excel(writer, sheet_name=f'Pesos_Locales_{criterio}', index=False)

    # 4) Exportar ranking final de alternativas
    df_ranking = df_resultados.copy()
    df_ranking = df_ranking.sort_values(by='Peso Global', ascending=False).reset_index(drop=True)
    df_ranking['Ranking'] = df_ranking.index + 1
    df_ranking.to_excel(writer, sheet_name='Ranking_Alternativas', index=False)

# Aplicar formato con openpyxl
wb = load_workbook(archivo_salida)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # encabezados en negrita y centrados
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    # formateo porcentual para valores [0,1]
    for col in ws.iter_cols(min_row=2, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                cell.number_format = '0.00%'
    # autoajuste de ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
wb.save(archivo_salida)

print(f"✅ Archivo Excel generado: {archivo_salida}")

# =========================
# Análisis de sensibilidad (gráfico radial)
# =========================
nombre_alternativas = alternativas
nombre_criterios = criterios
abreviaturas = {
    'Técnico': 'T',
    'Económico': 'E',
    'Social': 'S',
    'Ambiental': 'A',
    'Comercial': 'C'
}

# Generar combinaciones de criterios
combinaciones = []
etiquetas = []
for r in range(1, len(nombre_criterios) + 1):
    for combo in itertools.combinations(nombre_criterios, r):
        combinaciones.append(combo)
        etiquetas.append("-".join([abreviaturas.get(c, c[:1].upper()) for c in combo]))

# Asegurar formato de matriz local
matriz_local = np.array(pesos_locales)  # (criterios, alternativas)

# Recalcular pesos globales para cada combinación (ponderación equitativa dentro del subconjunto)
pesos_por_alternativa = np.zeros((len(nombre_alternativas), len(combinaciones)))
for idx, combo in enumerate(combinaciones):
    pesos_combo = np.zeros(len(nombre_criterios))
    for i, criterio in enumerate(nombre_criterios):
        if criterio in combo:
            pesos_combo[i] = 1 / len(combo)
    pesos_globales_combo = np.dot(matriz_local.T, pesos_combo)
    pesos_por_alternativa[:, idx] = pesos_globales_combo

# Exportar resultados a Excel
df_radial = pd.DataFrame(pesos_por_alternativa, index=nombre_alternativas, columns=etiquetas).T
df_radial.index.name = "Combinación de Criterios"
df_radial.reset_index(inplace=True)
archivo_excel_radial = "resultados_radiales_ahp.xlsx"
df_radial.to_excel(archivo_excel_radial, index=False)
print(f"✅ Resultados exportados a: {archivo_excel_radial}")

# === Graficar en coordenadas polares (todas las alternativas, estilo limpio) ===

# Usar todas las alternativas
k = len(nombre_alternativas)
alts_all = nombre_alternativas
datos_all = pesos_por_alternativa  # (k, n_combinaciones)

# Configurar ángulos
angles = np.linspace(0, 2 * np.pi, len(combinaciones), endpoint=False).tolist()
angles += angles[:1]

# Crear figura y ejes
plt.figure(figsize=(11, 11))
ax = plt.subplot(111, polar=True)

# Paleta de colores ampliada (tab20) para más diferenciación
colores = plt.cm.tab20(np.linspace(0, 1, k))

# Dibujar cada alternativa
for i, alt in enumerate(alts_all):
    datos = datos_all[i, :].tolist()
    datos += datos[:1]
    ax.plot(angles, datos, label=alt, linewidth=1.6, alpha=0.9, color=colores[i % len(colores)])
    ax.fill(angles, datos, alpha=0.18, color=colores[i % len(colores)])

# Escalado dinámico del eje radial
max_valor = datos_all.max()
ax.set_ylim(0, max_valor)
ax.set_yticks(np.linspace(0, max_valor, 5))
ax.set_yticklabels([f"{int(x * 100)}%" for x in np.linspace(0, max_valor, 5)], fontsize=9, color="gray")

# Ejes y estilo
ax.set_xticks(angles[:-1])
ax.set_xticklabels(etiquetas, fontsize=9)
ax.set_theta_offset(np.pi / 2)
ax.set_theta_direction(-1)
ax.grid(True, color="gray", linestyle="--", linewidth=0.6, alpha=0.6)

# Título y leyenda
ax.set_title("Análisis de Sensibilidad AHP – Todas las Alternativas", fontsize=14, pad=25, weight='bold')

# Leyenda limpia, de múltiples columnas si hay muchas alternativas
n_cols = 2 if k <= 10 else 3
ax.legend(
    loc='upper right',
    bbox_to_anchor=(1.35, 1.05),
    fontsize=8,
    frameon=False,
    ncol=n_cols,
    title="Alternativas",
    title_fontsize=9
)

# Ajustes finales
plt.tight_layout()
plt.savefig("grafico_radial_todas.png", dpi=300, bbox_inches='tight')
print("✅ Gráfico radial (todas las alternativas) guardado como 'grafico_radial_todas.png'")
plt.show()

# === Graficar en coordenadas polares (Top 10 alternativas, visual limpio) ===

# Seleccionar las 10 mejores alternativas (si hay menos, usa todas)
k = min(10, len(nombre_alternativas))
idx_top = np.argsort(pesos_globales)[::-1][:k]
alts_top = [nombre_alternativas[i] for i in idx_top]
datos_top = pesos_por_alternativa[idx_top, :]  # (k, n_combinaciones)

# Configurar ángulos
angles = np.linspace(0, 2 * np.pi, len(combinaciones), endpoint=False).tolist()
angles += angles[:1]

# Crear figura y ejes
plt.figure(figsize=(11, 11))
ax = plt.subplot(111, polar=True)

# Paleta de colores contrastantes (tab20)
colores = plt.cm.tab20(np.linspace(0, 1, k))

# Dibujar cada alternativa
for i, alt in enumerate(alts_top):
    datos = datos_top[i, :].tolist()
    datos += datos[:1]
    ax.plot(angles, datos, label=alt, linewidth=1.8, alpha=0.9, color=colores[i])
    ax.fill(angles, datos, alpha=0.18, color=colores[i])

# Escalado dinámico del eje radial
max_valor = datos_top.max()
ax.set_ylim(0, max_valor)
ax.set_yticks(np.linspace(0, max_valor, 5))
ax.set_yticklabels([f"{int(x * 100)}%" for x in np.linspace(0, max_valor, 5)], fontsize=9, color="gray")

# Ejes y estilo
ax.set_xticks(angles[:-1])
ax.set_xticklabels(etiquetas, fontsize=9)
ax.set_theta_offset(np.pi / 2)
ax.set_theta_direction(-1)
ax.grid(True, color="gray", linestyle="--", linewidth=0.6, alpha=0.6)

# Título y leyenda
ax.set_title("Análisis de Sensibilidad AHP – Top 10 Alternativas", fontsize=14, pad=25, weight='bold')

# Leyenda bien distribuida y legible
ax.legend(
    loc='upper right',
    bbox_to_anchor=(1.25, 1.05),
    fontsize=8,
    frameon=False,
    ncol=2,
    title="Alternativas",
    title_fontsize=9
)

# Ajustes finales
plt.tight_layout()
plt.savefig("grafico_radial_top10.png", dpi=300, bbox_inches='tight')
print("✅ Gráfico radial (Top 10) guardado como 'grafico_radial_top10.png'")
plt.show()

# === Graficar en coordenadas polares (Top 5 alternativas, estilo limpio) ===

# Seleccionar las 5 mejores alternativas (según peso global)
k = min(5, len(nombre_alternativas))
idx_top = np.argsort(pesos_globales)[::-1][:k]
alts_top = [nombre_alternativas[i] for i in idx_top]
datos_top = pesos_por_alternativa[idx_top, :]  # (k, n_combinaciones)

# Configurar ángulos
angles = np.linspace(0, 2 * np.pi, len(combinaciones), endpoint=False).tolist()
angles += angles[:1]

# Crear figura y ejes
plt.figure(figsize=(10, 10))
ax = plt.subplot(111, polar=True)

# Paleta de colores contrastantes (tab10)
colores = plt.cm.tab10(np.linspace(0, 1, k))

# Dibujar cada alternativa
for i, alt in enumerate(alts_top):
    datos = datos_top[i, :].tolist()
    datos += datos[:1]
    ax.plot(angles, datos, label=alt, linewidth=2.0, alpha=0.9, color=colores[i])
    ax.fill(angles, datos, alpha=0.25, color=colores[i])

# Escalado dinámico del eje radial
max_valor = datos_top.max()
ax.set_ylim(0, max_valor)
ax.set_yticks(np.linspace(0, max_valor, 5))
ax.set_yticklabels([f"{int(x * 100)}%" for x in np.linspace(0, max_valor, 5)], fontsize=9, color="gray")

# Ejes y estilo
ax.set_xticks(angles[:-1])
ax.set_xticklabels(etiquetas, fontsize=9)
ax.set_theta_offset(np.pi / 2)
ax.set_theta_direction(-1)
ax.grid(True, color="gray", linestyle="--", linewidth=0.6, alpha=0.6)

# Título y leyenda
ax.set_title("Análisis de Sensibilidad AHP – Top 5 Alternativas", fontsize=14, pad=25, weight='bold')

# Leyenda ordenada y legible
ax.legend(
    loc='upper right',
    bbox_to_anchor=(1.25, 1.05),
    fontsize=9,
    frameon=False,
    ncol=1,
    title="Alternativas",
    title_fontsize=10
)

# Ajustes finales
plt.tight_layout()
plt.savefig("grafico_radial_top5.png", dpi=300, bbox_inches='tight')
print("✅ Gráfico radial (Top 5) guardado como 'grafico_radial_top5.png'")
plt.show()